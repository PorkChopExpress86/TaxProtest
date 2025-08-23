from __future__ import annotations

from typing import List, Dict, Any

from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.http import HttpRequest, HttpResponse
from django.core.paginator import Paginator

from .forms import PropertySearchForm, ComparablesParamForm
from .services import find_comps

try:
    # Legacy search & extract functions (flat module at repo root)
    from extract_data import search_properties
except Exception:  # pragma: no cover - fallback
    search_properties = None  # type: ignore


SEARCH_PAGE_SIZES = [25, 50, 100, 200]


def _clean_int(val: str | None, default: int) -> int:
    try:
        return int(val) if val is not None else default
    except ValueError:
        return default


def search(request: HttpRequest) -> HttpResponse:
    """Render search form & results."""
    form = PropertySearchForm(request.GET or None)
    results: List[Dict[str, Any]] | None = None
    total = 0
    highlight_term = ""

    page_size = 50
    page_obj = None
    total_pages = 0

    if form.is_valid() and any(form.cleaned_data.get(f) for f in ["acct", "owner", "street", "zip_code"]):
        acct = form.cleaned_data.get("acct", "").strip()
        owner = form.cleaned_data.get("owner", "").strip()
        street = form.cleaned_data.get("street", "").strip()
        zip_code = form.cleaned_data.get("zip_code", "").strip()
        exact_match = bool(form.cleaned_data.get("exact_match"))
        highlight_term = owner or street
        if not search_properties:
            messages.error(request, "Search backend not available.")
        else:
            try:
                all_results = search_properties(acct, street, zip_code, owner, exact_match)  # type: ignore
                total = len(all_results)
                if total == 0:
                    messages.info(request, "No properties found. Try partial names or different spelling.")
                else:
                    messages.success(request, f"Found {total} matching properties.")
                page_size = int(form.cleaned_data.get("page_size") or 50)
                if page_size not in SEARCH_PAGE_SIZES:
                    page_size = 50
                paginator = Paginator(all_results, page_size)
                page_number = request.GET.get("page") or "1"
                try:
                    page_obj = paginator.get_page(page_number)
                except Exception:
                    page_obj = paginator.get_page(1)
                results = list(page_obj.object_list)  # copy for template iteration
                total_pages = paginator.num_pages
            except Exception as e:  # pragma: no cover - defensive
                messages.error(request, f"Error during search: {e}")
    elif request.GET:
        # Submitted but invalid / empty
        if form.is_valid():  # no criteria provided
            messages.warning(request, "Enter at least one search criterion.")

    context = {
        "form": form,
        "results": results,
        "total": total,
        "highlight_term": highlight_term,
        "page_obj": page_obj,
        "page_size": page_size,
        "total_pages": total_pages,
    }
    return render(request, "search.html", context)


def export_search_results(request: HttpRequest, fmt: str) -> HttpResponse:
    """Export search results as CSV or XLSX."""
    form = PropertySearchForm(request.GET or None)
    
    if not (form.is_valid() and any(form.cleaned_data.get(f) for f in ["acct", "owner", "street", "zip_code"])):
        return HttpResponse("Invalid search parameters", status=400)
    
    # Re-run search to get results
    acct = form.cleaned_data.get("acct", "").strip()
    owner = form.cleaned_data.get("owner", "").strip()
    street = form.cleaned_data.get("street", "").strip()
    zip_code = form.cleaned_data.get("zip_code", "").strip()
    exact_match = bool(form.cleaned_data.get("exact_match"))
    
    if not search_properties:
        return HttpResponse("Search backend not available", status=500)
    
    try:
        results = search_properties(acct, street, zip_code, owner, exact_match)  # type: ignore
    except Exception as e:
        return HttpResponse(f"Error during search: {e}", status=500)
    
    if fmt == 'xlsx':
        # XLSX export using openpyxl
        from openpyxl import Workbook
        import pandas as pd
        from io import BytesIO
        
        df = pd.DataFrame(results) if results else pd.DataFrame([{"acct": "no results", "note": "no properties found"}])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"
        
        # Write dataframe to worksheet
        for r_idx, row in enumerate(df.itertuples(index=False), 1):
            if r_idx == 1:
                # Write header
                for c_idx, col in enumerate(df.columns, 1):
                    ws.cell(row=r_idx, column=c_idx, value=col)
                r_idx += 1
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        filename = "search_results.xlsx"
        resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f"attachment; filename=\"{filename}\""
        return resp
    else:
        # CSV export
        import csv
        from io import StringIO

        buf = StringIO()
        writer = csv.writer(buf)
        if results:
            header = list(results[0].keys())
            writer.writerow(header)
            for r in results:
                writer.writerow([r.get(k, "") for k in header])
        else:
            writer.writerow(["note"])
            writer.writerow(["no properties found"])

        csv_data = buf.getvalue()
        buf.close()

        filename = "search_results.csv"
        resp = HttpResponse(csv_data, content_type="text/csv")
        resp["Content-Disposition"] = f"attachment; filename=\"{filename}\""
        return resp


def export_comparables(request: HttpRequest, acct: str, fmt: str) -> HttpResponse:
    """Export comparables as CSV or XLSX."""
    # Read params
    max_c = _clean_int(request.GET.get("max"), 25)
    min_c = _clean_int(request.GET.get("min"), 20)
    max_radius = request.GET.get("max_radius")
    strict_first = bool(request.GET.get("strict_first"))
    try:
        result = find_comps(acct, max_comps=max_c, min_comps=min_c, radius_first_strict=strict_first, max_radius=max_radius)
        comps = result.get("comps", [])
        subject = result.get("subject")
    except Exception as e:
        return HttpResponse(f"Error computing comparables: {e}", status=500)

    if fmt == 'xlsx':
        # XLSX export using openpyxl
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparables"
        
        if comps:
            df = pd.DataFrame(comps)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
        else:
            ws.append(["acct", "note"])
            ws.append([acct, "no comparables"])
            
        # Add subject info if available
        if subject:
            ws2 = wb.create_sheet("Subject")
            ws2.append(["Property", "Value"])
            for k, v in subject.items():
                ws2.append([k, v])
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        filename = f"comparables_{acct}.xlsx"
        resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f"attachment; filename=\"{filename}\""
        return resp
    else:
        # CSV export
        import csv
        from io import StringIO

        buf = StringIO()
        writer = csv.writer(buf)
        if comps:
            # header from keys of first comp
            header = list(comps[0].keys())
            writer.writerow(header)
            for c in comps:
                writer.writerow([c.get(k, "") for k in header])
        else:
            writer.writerow(["acct", "note"])
            writer.writerow([acct, "no comparables"])

        csv_data = buf.getvalue()
        buf.close()

        filename = f"comparables_{acct}.csv"
        resp = HttpResponse(csv_data, content_type="text/csv")
        resp["Content-Disposition"] = f"attachment; filename=\"{filename}\""
        return resp


def comparables_view(request: HttpRequest, acct: str) -> HttpResponse:
    form = ComparablesParamForm(request.GET or None)
    comps: List[Dict[str, Any]] = []
    subject: Dict[str, Any] | None = None
    meta: Dict[str, Any] = {}
    if form.is_valid():
        max_c = form.cleaned_data.get("max") or 25
        min_c = form.cleaned_data.get("min") or 20
        if min_c > max_c:
            min_c = max_c
        max_radius = form.cleaned_data.get("max_radius")
        strict_first = bool(form.cleaned_data.get("strict_first"))
        try:
            result = find_comps(
                acct,
                max_comps=max_c,
                min_comps=min_c,
                radius_first_strict=strict_first,
                max_radius=max_radius,
            )
            subject = result.get("subject")
            comps = result.get("comps", [])
            meta = result.get("meta", {})
            if not comps:
                messages.info(request, "No comparables found with current parameters.")
        except Exception as e:  # pragma: no cover - defensive
            messages.error(request, f"Error computing comparables: {e}")
    else:
        form = ComparablesParamForm(initial={"max": 25, "min": 20})

    return render(
        request,
        "comparables.html",
        {
            "acct": acct,
            "form": form,
            "subject": subject,
            "comparables": comps,
            "meta": meta,
        },
    )
